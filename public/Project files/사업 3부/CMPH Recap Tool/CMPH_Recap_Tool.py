import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, os, io, copy, re, datetime
from PIL import Image as PILImage, ImageTk
import fitz
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

BG     = "#f5f5f0"
CARD   = "#ffffff"
BORDER = "#d0d0c8"
ACCENT = "#2563eb"
TEXT   = "#1a1a1a"
MUTED  = "#6b7280"
BTN_BG = "#2563eb"
BTN_FG = "#ffffff"

YELLOW_FILL  = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
NO_FILL      = PatternFill(fill_type=None)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGN = Alignment(wrap_text=True, vertical='center', horizontal='center')

# ── PDF helpers ────────────────────────────────────────────────────────────────

def parse_pdf_text(pdf_path):
    doc = fitz.open(pdf_path)
    full = ""
    for page in doc:
        full += page.get_text() + "\n"
    r = {}
    m = re.search(r'Product\s+Number\s*[:\-]?\s*(\d{7})', full, re.I)
    if m: r['style'] = m.group(1)
    m = re.search(r'Product\s+Name\s*[:\-]?\s*(.+?)(?:Actual\s+Size|Season|Department)', full, re.I|re.S)
    if m:
        name = re.sub(r'\s+', ' ', m.group(1)).strip()
        name = re.sub(r'\s*F\s*\d{7}.*$', '', name).strip()  # 끝의 'F 1234567...' 제거
        r['name'] = name[:120]
    m = re.search(r'Main\s+Fabric\s+Description\s*[:\-]?\s*(.+?)(?:Main\s+Fabric\s+Supplier|Vendor\s*:|Factory\s*:|Country)', full, re.I|re.S)
    if m: r['fabric'] = re.sub(r'\s+', ' ', m.group(1)).strip()
    nl = r.get('name', '').lower()
    r['category'] = 'LIVI' if 'livi' in nl else 'BOTTOM' if ('pant' in nl or 'jogger' in nl) else 'JACKET' if 'jacket' in nl else 'KNIT TOP'
    return r


def extract_sketch(pdf_path):
    """커버 페이지(page 0)의 첫 번째 가로 이미지 사용 (노란 동그라미 스케치)"""
    doc = fitz.open(pdf_path)
    # page 0 우선, 없으면 전체에서 탐색
    for pnum in [0, 1]:
        if pnum >= doc.page_count:
            continue
        imgs = doc[pnum].get_images(full=True)
        for img_info in imgs:
            xref = img_info[0]
            pix = fitz.Pixmap(doc, xref)
            if pix.width > pix.height:  # 가로형
                return PILImage.open(io.BytesIO(pix.tobytes("png")))
    # fallback: 가장 큰 가로형
    best_pix, best_area = None, 0
    for pnum in range(doc.page_count):
        for img_info in doc[pnum].get_images(full=True):
            xref = img_info[0]
            pix = fitz.Pixmap(doc, xref)
            if pix.width > pix.height and pix.width * pix.height > best_area:
                best_area = pix.width * pix.height
                best_pix = pix
    return PILImage.open(io.BytesIO(best_pix.tobytes("png"))) if best_pix else None


def find_last_subtotal_row(ws):
    last = 3
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and 'SUBTOTAL' in str(val):
            last = r
    return last


def copy_merge_pattern(ws, ref_row, new_row):
    offset = new_row - ref_row
    to_add = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == ref_row:
            to_add.append((mr.min_row + offset, mr.min_col, mr.max_row + offset, mr.max_col))
    for (r1, c1, r2, c2) in to_add:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def add_row_to_excel(excel_path, pdf_path, form, out_path, log_fn=None):
    def log(msg):
        if log_fn: log_fn(msg)

    wb = load_workbook(excel_path)
    ws = wb.active

    ref_row = find_last_subtotal_row(ws)
    new_row = ref_row + 4
    sub_row = new_row + 2

    today_dt  = datetime.datetime.now()
    today_str = today_dt.strftime('%Y-%m-%d %H:%M')

    season_str = form.get('season', "HOL'26")
    season_div = f"{season_str}\n{form.get('category','KNIT TOP')}\n{form.get('assign','Hansoll assigned')}"

    # ── 스타일 복사 헬퍼 (테두리 포함, fill 초기화) ───────────────────────────
    def copy_cell_style(src, dst):
        if src.font:   dst.font   = copy.copy(src.font)
        if src.border: dst.border = copy.copy(src.border)
        dst.number_format = src.number_format
        dst.fill = NO_FILL

    # ── 4행 블록 전체 스타일 복사 (테두리 보존) ───────────────────────────────
    for block_offset in range(4):
        src_r = ref_row + block_offset
        dst_r = new_row + block_offset
        ws.row_dimensions[dst_r].height = ws.row_dimensions[src_r].height or 29.0
        for col_idx in range(1, 15):
            try:
                copy_cell_style(ws.cell(row=src_r, column=col_idx),
                                ws.cell(row=dst_r, column=col_idx))
            except: pass

    # ── 머지 패턴 복사 ────────────────────────────────────────────────────────
    copy_merge_pattern(ws, ref_row, new_row)
    log(f"✓ 스타일/머지 복사 완료")

    # ── 노란 하이라이트 (4행 전체, 루트셀만) ───────────────────────────────────
    for r in range(new_row, new_row + 4):
        for col_idx in range(1, 15):
            c = ws.cell(row=r, column=col_idx)
            # MergedCell은 루트셀이 아니므로 스킵 (루트셀에 이미 적용됨)
            if type(c).__name__ != 'MergedCell':
                c.fill = YELLOW_FILL

    # ── 데이터 입력 ───────────────────────────────────────────────────────────
    ws[f'A{new_row}'] = f'=SUBTOTAL(103,D$4:D{new_row})'
    ws[f'A{new_row}'].alignment = CENTER_ALIGN

    ws[f'B{new_row}'] = today_dt
    ws[f'B{new_row}'].number_format = 'm"/"d;@'   # 기존 형식과 동일
    ws[f'B{new_row}'].alignment = CENTER_ALIGN

    ws[f'C{new_row}'] = season_div
    ws[f'C{new_row}'].alignment = CENTER_ALIGN

    if form.get('style'):
        ws[f'D{new_row}'] = int(form['style'])
        ws[f'D{new_row}'].alignment = CENTER_ALIGN

    # E열(스케치)은 이미지로 처리
    ws[f'F{new_row}'] = form.get('name', '')
    ws[f'F{new_row}'].alignment = CENTER_ALIGN

    ws[f'G{new_row}'] = form.get('fabric', '')
    ws[f'G{new_row}'].alignment = Alignment(wrap_text=True, vertical='center')

    if form.get('cmph_k'):
        ws[f'K{new_row}'] = float(form['cmph_k'])
        ws[f'K{new_row}'].alignment = CENTER_ALIGN
    if form.get('cmph_l'):
        ws[f'L{new_row}'] = float(form['cmph_l'])
        ws[f'L{new_row}'].alignment = CENTER_ALIGN
    if form.get('cmph_m'):
        ws[f'M{new_row}'] = float(form['cmph_m'])
        ws[f'M{new_row}'].alignment = CENTER_ALIGN

    ws[f'N{new_row}'] = f"★ ADDED {today_str}"
    ws[f'N{new_row}'].alignment = Alignment(wrap_text=True, vertical='center')

    log(f"✓ Row {new_row} 추가 (S# {form.get('style','')})")

    # ── 서브행 ────────────────────────────────────────────────────────────────
    if not form.get('sub_enabled'):
        # sub 없을 때: +2/+3 독립셀의 top/bottom 테두리 제거 → 줄 사라짐, left/right 유지
        from openpyxl.styles import Border, Side
        for r_off in [2, 3]:
            for col_idx in range(1, 15):
                c = ws.cell(row=new_row + r_off, column=col_idx)
                if type(c).__name__ != 'MergedCell':
                    b = c.border
                    c.border = Border(
                        left=copy.copy(b.left) if b.left else Side(),
                        right=copy.copy(b.right) if b.right else Side(),
                        top=Side(),
                        bottom=Side()
                    )

    if form.get('sub_enabled'):
        # sub_row: H(qty), L(price), M(cmph) 만 쓸 수 있음 (나머지는 머지 내부)
        ws[f'H{sub_row}'] = form.get('sub_qty', '')
        ws[f'H{sub_row}'].alignment = CENTER_ALIGN
        if form.get('sub_cmph'):
            ws[f'M{sub_row}'] = float(form['sub_cmph'])
            ws[f'M{sub_row}'].alignment = CENTER_ALIGN
        if form.get('sub_l'):
            ws[f'L{sub_row}'] = float(form['sub_l'])
            ws[f'L{sub_row}'].alignment = CENTER_ALIGN
        log(f"✓ 서브행 Row {sub_row} 추가")



    # ── 스케치 이미지 ─────────────────────────────────────────────────────────
    sketch_pil = extract_sketch(pdf_path)
    if sketch_pil:
        sketch_resized = sketch_pil.resize((172, 100), PILImage.LANCZOS)
        buf = io.BytesIO()
        sketch_resized.save(buf, format='PNG')
        buf.seek(0)
        xl_img = XLImage(buf)
        xl_img.width  = 172
        xl_img.height = 100
        row_0 = new_row - 1
        anchor = TwoCellAnchor()
        anchor.editAs = 'oneCell'
        anchor._from  = AnchorMarker(col=4, row=row_0,     colOff=916782,  rowOff=88627)
        anchor.to     = AnchorMarker(col=4, row=row_0 + 3, colOff=2782888, rowOff=325711)
        xl_img.anchor = anchor
        ws.add_image(xl_img)
        log("✓ 스케치 삽입")
    else:
        log("⚠ 스케치 없음")

    wb.save(out_path)
    log(f"✓ 저장 완료: {os.path.basename(out_path)}")
    return new_row


# ── GUI ────────────────────────────────────────────────────────────────────────

SEASONS   = ["SPRING", "SUMMER", "FALL", "HOLIDAY"]
YEARS     = [str(y) for y in range(26, 33)]   # 26 ~ 32

class CMPHApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CMPH Recap — PDF → Excel")
        self.geometry("760x920")
        self.resizable(True, True)
        self.configure(bg=BG)
        self.excel_path = None
        self.pdf_paths  = []
        self.sketch_img = None
        self._build()

    def _build(self):
        canvas = tk.Canvas(self, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self.frame = tk.Frame(canvas, bg=BG)
        fid = canvas.create_window((0, 0), window=self.frame, anchor="nw")
        self.frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(fid, width=e.width))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # ── Header ────────────────────────────────────────────────────────────
        tk.Label(self.frame, text="CMPH RECAP  PDF → Excel",
                 font=("Helvetica",18,"bold"), bg=BG, fg=ACCENT).pack(anchor="w", padx=24, pady=(24,2))
        tk.Label(self.frame, text="택팩 PDF → 자동 파싱 → 기존 엑셀 양식에 행 추가  |  PDF 여러 개 동시 추가 가능",
                 font=("Helvetica",10), bg=BG, fg=MUTED).pack(anchor="w", padx=24, pady=(0,16))
        self._sep()

        # ── Upload ────────────────────────────────────────────────────────────
        up = tk.Frame(self.frame, bg=BG)
        up.pack(fill="x", padx=24, pady=6)
        self._dz(up, "📊  CMPH Recap (.xlsx)", self._pick_excel, "xl_lbl").pack(side="left", expand=True, fill="x", padx=(0,8))
        self._dz(up, "📄  Tech Pack PDF (여러 개 가능)", self._pick_pdfs, "pdf_lbl").pack(side="left", expand=True, fill="x")
        self._sep()

        # ── PDF list ──────────────────────────────────────────────────────────
        tk.Label(self.frame, text="선택된 PDF 목록",
                 font=("Helvetica",9,"bold"), bg=BG, fg=MUTED).pack(anchor="w", padx=24)
        self.pdf_list_frame = tk.Frame(self.frame, bg=CARD)
        self.pdf_list_frame.pack(fill="x", padx=24, pady=4)
        tk.Label(self.pdf_list_frame, text="PDF 없음",
                 font=("Helvetica",10), bg=CARD, fg=MUTED).pack(padx=12, pady=8)
        self._sep()

        # ── Sketch preview ────────────────────────────────────────────────────
        tk.Label(self.frame, text="스케치 미리보기 (첫 번째 PDF)",
                 font=("Helvetica",9,"bold"), bg=BG, fg=MUTED).pack(anchor="w", padx=24)
        sketch_outer = tk.Frame(self.frame, bg=CARD)
        sketch_outer.pack(fill="x", padx=24, pady=4)
        self.sketch_label = tk.Label(sketch_outer, bg=CARD, text="—", fg=MUTED, font=("Helvetica",10))
        self.sketch_label.pack(pady=12)
        self._sep()

        # ── Parsed info ───────────────────────────────────────────────────────
        tk.Label(self.frame, text="파싱된 정보 (수정 가능)",
                 font=("Helvetica",9,"bold"), bg=BG, fg=MUTED).pack(anchor="w", padx=24, pady=(4,2))

        self.f_style  = self._field("S/#")
        self.f_name   = self._field("PROGRAM NAME", wide=True)
        self.f_fabric = self._field("FABRIC DESCRIPTION", wide=True, multiline=True)

        # Season 선택 (시즌명 + 연도)
        season_row = tk.Frame(self.frame, bg=BG)
        season_row.pack(fill="x", padx=24, pady=3)

        tk.Label(season_row, text="Season", font=("Helvetica",9), bg=BG, fg=MUTED).pack(anchor="w")
        season_inner = tk.Frame(season_row, bg=BG)
        season_inner.pack(fill="x")

        self.f_season_name = self._inline_combo(season_inner, "", SEASONS)
        self.f_season_name.set_value("HOLIDAY")
        tk.Label(season_inner, text="'", font=("Helvetica",14,"bold"), bg=BG, fg=TEXT).pack(side="left", pady=2)
        self.f_season_year = self._inline_combo(season_inner, "", YEARS)
        self.f_season_year.set_value("26")

        # 시즌 미리보기
        self.season_preview = tk.Label(season_row, text="HOL'26",
                                        font=("Courier",11,"bold"), bg=BG, fg=ACCENT)
        self.season_preview.pack(anchor="w", pady=(2,0))

        # 변경 시 미리보기 업데이트
        def update_preview(*_):
            s = self.f_season_name.get_value()
            y = self.f_season_year.get_value()
            abbr = {'SPRING':'SPR','SUMMER':'SUM','FALL':'FAL','HOLIDAY':'HOL'}.get(s, s[:3])
            self.season_preview.config(text=f"{abbr}'{y}")
        self.f_season_name._cb_var.trace_add('write', update_preview)
        self.f_season_year._cb_var.trace_add('write', update_preview)

        # Category / Assignment
        cat_row = tk.Frame(self.frame, bg=BG)
        cat_row.pack(fill="x", padx=24, pady=3)
        self.f_cat    = self._inline_combo(cat_row, "Category",   ["KNIT TOP","LIVI","BOTTOM","JACKET"])
        self.f_assign = self._inline_combo(cat_row, "Assignment", ["Hansoll assigned","dual costing"])

        self._sep()

        # ── CMPH ──────────────────────────────────────────────────────────────
        tk.Label(self.frame, text="💰 CMPH 값",
                 font=("Helvetica",9,"bold"), bg=BG, fg=MUTED).pack(anchor="w", padx=24, pady=(4,2))
        cmph_row = tk.Frame(self.frame, bg=BG)
        cmph_row.pack(fill="x", padx=24, pady=4)
        self.f_k = self._cmph(cmph_row, "Hansoll Target\nCMPH  (K열)")
        self.f_l = self._cmph(cmph_row, "TSHY CMPH\nT/P 기준  (L열)")
        self.f_m = self._cmph(cmph_row, "TSHY CMPH\nActual  (M열)")
        self._sep()

        # ── Sub-row ───────────────────────────────────────────────────────────
        self.sub_var = tk.BooleanVar()
        tk.Checkbutton(self.frame, text=" 코스트 옵션 서브행 추가",
                       variable=self.sub_var, command=self._toggle_sub,
                       bg=BG, fg=TEXT, selectcolor=BG, activebackground=BG,
                       font=("Helvetica",11)).pack(anchor="w", padx=24, pady=(6,2))
        self.sub_fr = tk.Frame(self.frame, bg=CARD)
        self.f_sub_qty    = self._field("Q'TY / 설명", multiline=True, parent=self.sub_fr)
        self.f_sub_l      = self._field("서브행 단가 (L열)", parent=self.sub_fr)
        self.f_sub_cmph   = self._field("서브행 CMPH (M열)", parent=self.sub_fr)
        self._sep()

        # ── Log ───────────────────────────────────────────────────────────────
        tk.Label(self.frame, text="로그",
                 font=("Helvetica",9,"bold"), bg=BG, fg=MUTED).pack(anchor="w", padx=24)
        log_fr = tk.Frame(self.frame, bg=CARD)
        log_fr.pack(fill="x", padx=24, pady=4)
        self.log_box = tk.Text(log_fr, height=5, bg=CARD, fg=TEXT,
                               font=("Courier",10), relief="flat",
                               state="disabled", wrap="word")
        self.log_box.pack(fill="x", padx=8, pady=8)
        self._sep()

        # ── Buttons ───────────────────────────────────────────────────────────
        btn_fr = tk.Frame(self.frame, bg=BG)
        btn_fr.pack(fill="x", padx=24, pady=(6,32))
        tk.Button(btn_fr, text="＋  행 추가 & 저장",
                  font=("Helvetica",13,"bold"), bg=BTN_BG, fg=BTN_FG,
                  relief="flat", padx=24, pady=12, cursor="hand2",
                  command=self._run).pack(side="left", padx=(0,12))
        tk.Button(btn_fr, text="초기화",
                  font=("Helvetica",11), bg=BORDER, fg=TEXT,
                  relief="flat", padx=16, pady=12, cursor="hand2",
                  command=self._reset).pack(side="left")

    # ── Widget helpers ─────────────────────────────────────────────────────────

    def _sep(self):
        tk.Frame(self.frame, bg=BORDER, height=1).pack(fill="x", padx=24, pady=8)

    def _dz(self, parent, label, cmd, attr):
        fr = tk.Frame(parent, bg=CARD, cursor="hand2")
        fr.bind("<Button-1>", lambda e: cmd())
        tk.Label(fr, text=label, font=("Helvetica",10,"bold"), bg=CARD, fg=TEXT).pack(pady=(12,2))
        lbl = tk.Label(fr, text="클릭하여 선택", font=("Helvetica",9), bg=CARD, fg=MUTED, wraplength=220)
        lbl.pack(pady=(0,12))
        setattr(self, attr, lbl)
        return fr

    def _field(self, label, wide=False, multiline=False, parent=None):
        p = parent or self.frame
        fr = tk.Frame(p, bg=p.cget('bg'))
        fr.pack(fill="x", padx=24, pady=3)
        tk.Label(fr, text=label, font=("Helvetica",9), bg=p.cget('bg'), fg=MUTED).pack(anchor="w")
        if multiline:
            w = tk.Text(fr, height=3, font=("Courier",10), bg=CARD, fg=TEXT,
                        relief="flat", insertbackground="#1a1a1a", wrap="word")
            w.pack(fill="x", ipady=4, pady=2)
            w.get_value = lambda: w.get("1.0","end-1c")
            w.set_value = lambda v: (w.delete("1.0","end"), w.insert("1.0",v))
        else:
            v = tk.StringVar()
            w = tk.Entry(fr, textvariable=v, font=("Courier",10), bg=CARD, fg=TEXT,
                         relief="flat", insertbackground="#1a1a1a")
            w.pack(fill="x", ipady=6, pady=2)
            w.get_value = lambda: v.get()
            w.set_value = lambda x: v.set(x)
        return w

    def _inline_combo(self, parent, label, values):
        fr = tk.Frame(parent, bg=BG)
        fr.pack(side="left", expand=True, fill="x", padx=(0,8))
        if label:
            tk.Label(fr, text=label, font=("Helvetica",9), bg=BG, fg=MUTED).pack(anchor="w")
        v = tk.StringVar(value=values[0])
        cb = ttk.Combobox(fr, textvariable=v, values=values, state="readonly",
                          font=("Courier",10), width=12)
        cb.pack(fill="x")
        cb.get_value = lambda: v.get()
        cb.set_value = lambda x: v.set(x)
        cb._cb_var   = v   # trace용
        return cb

    def _cmph(self, parent, label):
        fr = tk.Frame(parent, bg=BG)
        fr.pack(side="left", expand=True, fill="x", padx=4)
        tk.Label(fr, text=label, font=("Helvetica",9), bg=BG, fg=MUTED, justify="center").pack()
        v = tk.StringVar()
        e = tk.Entry(fr, textvariable=v, font=("Courier",12,"bold"), bg=CARD,
                     fg=ACCENT, relief="flat", justify="center", insertbackground="#2563eb")
        e.pack(fill="x", ipady=8, pady=2)
        e.get_value = lambda: v.get()
        e.set_value = lambda x: v.set(x)
        return e

    def _get_season_str(self):
        s = self.f_season_name.get_value()
        y = self.f_season_year.get_value()
        abbr = {'SPRING':'SPR','SUMMER':'SUM','FALL':'FAL','HOLIDAY':'HOL'}.get(s, s[:3])
        return f"{abbr}'{y}"

    # ── Actions ────────────────────────────────────────────────────────────────

    def _pick_excel(self):
        p = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if p:
            self.excel_path = p
            self.xl_lbl.config(text=os.path.basename(p), fg=ACCENT)
            self._log(f"Excel: {os.path.basename(p)}")

    def _pick_pdfs(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF","*.pdf")])
        if not paths: return
        self.pdf_paths = list(paths)
        self.pdf_lbl.config(text=f"{len(paths)}개 선택됨", fg=ACCENT)
        self._update_pdf_list()
        threading.Thread(target=self._parse_first_pdf, daemon=True).start()

    def _update_pdf_list(self):
        for w in self.pdf_list_frame.winfo_children():
            w.destroy()
        if not self.pdf_paths:
            tk.Label(self.pdf_list_frame, text="PDF 없음",
                     font=("Helvetica",10), bg=CARD, fg=MUTED).pack(padx=12, pady=8)
            return
        for i, p in enumerate(self.pdf_paths):
            row = tk.Frame(self.pdf_list_frame, bg=CARD)
            row.pack(fill="x", padx=8, pady=2)
            tk.Label(row, text=f"{i+1}.", font=("Courier",10), bg=CARD, fg=MUTED, width=3).pack(side="left")
            tk.Label(row, text=os.path.basename(p), font=("Courier",10), bg=CARD, fg=TEXT).pack(side="left")

    def _parse_first_pdf(self):
        if not self.pdf_paths: return
        pdf = self.pdf_paths[0]
        try:
            self._log(f"파싱 중: {os.path.basename(pdf)}")
            d = parse_pdf_text(pdf)
            def update():
                if d.get('style'):    self.f_style.set_value(d['style'])
                if d.get('name'):     self.f_name.set_value(d['name'])
                if d.get('fabric'):   self.f_fabric.set_value(d['fabric'])
                if d.get('category'): self.f_cat.set_value(d['category'])
                self._log(f"✓ 파싱 완료 — S# {d.get('style','?')}")
            self.after(0, update)
            sketch = extract_sketch(pdf)
            if sketch:
                thumb = sketch.resize((300, int(300 * sketch.height / sketch.width)), PILImage.LANCZOS)
                photo = ImageTk.PhotoImage(thumb)
                def show():
                    self.sketch_img = photo
                    self.sketch_label.config(image=photo, text="")
                self.after(0, show)
                self._log("✓ 스케치 추출됨")
        except Exception as e:
            self._log(f"오류: {e}")

    def _toggle_sub(self):
        if self.sub_var.get():
            self.sub_fr.pack(fill="x", padx=24, pady=4)
        else:
            self.sub_fr.pack_forget()

    def _run(self):
        if not self.excel_path:
            messagebox.showerror("오류", "Excel 파일을 선택하세요"); return
        if not self.pdf_paths:
            messagebox.showerror("오류", "PDF 파일을 선택하세요"); return
        out = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=os.path.basename(self.excel_path).replace('.xlsx','_updated.xlsx'),
            filetypes=[("Excel","*.xlsx")])
        if not out: return
        form = {
            'style':       self.f_style.get_value(),
            'name':        self.f_name.get_value(),
            'fabric':      self.f_fabric.get_value(),
            'season':      self._get_season_str(),
            'category':    self.f_cat.get_value(),
            'assign':      self.f_assign.get_value(),
            'cmph_k':      self.f_k.get_value(),
            'cmph_l':      self.f_l.get_value(),
            'cmph_m':      self.f_m.get_value(),
            'sub_enabled': self.sub_var.get(),
            'sub_qty':     self.f_sub_qty.get_value(),
            'sub_cmph':    self.f_sub_cmph.get_value(),
            'sub_l':       self.f_sub_l.get_value(),
        }
        threading.Thread(target=self._do_run, args=(form, out), daemon=True).start()

    def _do_run(self, form, out_path):
        try:
            self._log(f"처리 시작 — PDF {len(self.pdf_paths)}개")
            for i, pdf_path in enumerate(self.pdf_paths):
                self._log(f"[{i+1}/{len(self.pdf_paths)}] {os.path.basename(pdf_path)}")
                src = self.excel_path if i == 0 else out_path
                if i == 0:
                    row_form = dict(form)
                else:
                    d = parse_pdf_text(pdf_path)
                    row_form = {
                        'style': d.get('style',''), 'name': d.get('name',''),
                        'fabric': d.get('fabric',''),
                        'season': form['season'],
                        'category': d.get('category','KNIT TOP'),
                        'assign': form.get('assign','Hansoll assigned'),
                        'cmph_k':'', 'cmph_l':'', 'cmph_m':'',
                        'sub_enabled': False,
                        'sub_qty':'', 'sub_cmph':'', 'sub_remark':'',
                    }
                add_row_to_excel(src, pdf_path, row_form, out_path, self._log)
            def done():
                messagebox.showinfo("완료!", f"{len(self.pdf_paths)}개 처리 완료\n\n저장: {out_path}")
                if os.name == 'nt':
                    os.startfile(out_path)
                else:
                    os.system(f'open "{out_path}"')
            self.after(0, done)
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("오류", str(e)))
            self._log(f"오류: {e}")

    def _reset(self):
        self.excel_path = None
        self.pdf_paths  = []
        self.sketch_img = None
        self.xl_lbl.config(text="클릭하여 선택", fg=MUTED)
        self.pdf_lbl.config(text="클릭하여 선택", fg=MUTED)
        self.sketch_label.config(image="", text="—")
        self._update_pdf_list()
        for w in [self.f_style, self.f_name, self.f_fabric,
                  self.f_k, self.f_l, self.f_m,
                  self.f_sub_qty, self.f_sub_l, self.f_sub_cmph]:
            w.set_value("")
        self.f_season_name.set_value("HOLIDAY")
        self.f_season_year.set_value("26")
        self.f_cat.set_value("KNIT TOP")
        self.f_assign.set_value("Hansoll assigned")
        self.sub_var.set(False)
        self.sub_fr.pack_forget()
        self.log_box.config(state="normal")
        self.log_box.delete("1.0","end")
        self.log_box.config(state="disabled")

    def _log(self, msg):
        def _do():
            ts = datetime.datetime.now().strftime('%H:%M:%S')
            self.log_box.config(state="normal")
            self.log_box.insert("end", f"[{ts}] {msg}\n")
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.after(0, _do)


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) > 1 and sys.argv[1] == "--parse":
        try:
            d = parse_pdf_text(sys.argv[2])
            print(json.dumps(d))
        except Exception as e:
            print(json.dumps({"error": str(e)}))
        sys.exit(0)
    elif len(sys.argv) > 1 and sys.argv[1] == "--generate":
        try:
            args = sys.argv[2:]
            excel_path = args[args.index("--excel") + 1]
            out_path = args[args.index("--out") + 1]
            form_json = args[args.index("--form") + 1]
            form = json.loads(form_json)
            
            pdfs = []
            for i, arg in enumerate(args):
                if arg == "--pdf":
                    pdfs.append(args[i+1])
            
            for i, pdf in enumerate(pdfs):
                src = excel_path if i == 0 else out_path
                if i == 0:
                    row_form = dict(form)
                else:
                    d = parse_pdf_text(pdf)
                    row_form = {
                        'style': d.get('style',''), 'name': d.get('name',''),
                        'fabric': d.get('fabric',''),
                        'season': form.get('season',''),
                        'category': d.get('category','KNIT TOP'),
                        'assign': form.get('assign','Hansoll assigned'),
                        'cmph_k':'', 'cmph_l':'', 'cmph_m':'',
                        'sub_enabled': False,
                        'sub_qty':'', 'sub_cmph':'', 'sub_remark':''
                    }
                add_row_to_excel(src, pdf, row_form, out_path)
            print(json.dumps({"success": True, "out_path": out_path}))
        except Exception as e:
            print(json.dumps({"error": str(e)}))
        sys.exit(0)
    else:
        app = CMPHApp()
        app.mainloop()

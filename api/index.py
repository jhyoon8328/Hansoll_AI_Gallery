from flask import Flask, request, jsonify, send_file
import io, re, datetime, os
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
import copy
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

app = Flask(__name__)

YELLOW_FILL  = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
NO_FILL      = PatternFill(fill_type=None)
CENTER_ALIGN = Alignment(wrap_text=True, vertical='center', horizontal='center')

def parse_pdf_text_from_bytes(pdf_bytes):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    full = ""
    for page in doc:
        full += page.get_text() + "\n"
    r = {}
    m = re.search(r'Product\s+Number\s*[:\-]?\s*(\d{7})', full, re.I)
    if m: r['style'] = m.group(1)
    m = re.search(r'Product\s+Name\s*[:\-]?\s*(.+?)(?:Actual\s+Size|Season|Department)', full, re.I|re.S)
    if m:
        name = re.sub(r'\s+', ' ', m.group(1)).strip()
        name = re.sub(r'\s*F\s*\d{7}.*$', '', name).strip()
        r['name'] = name[:120]
    m = re.search(r'Main\s+Fabric\s+Description\s*[:\-]?\s*(.+?)(?:Main\s+Fabric\s+Supplier|Vendor\s*:|Factory\s*:|Country)', full, re.I|re.S)
    if m: r['fabric'] = re.sub(r'\s+', ' ', m.group(1)).strip()
    nl = r.get('name', '').lower()
    r['category'] = 'LIVI' if 'livi' in nl else 'BOTTOM' if ('pant' in nl or 'jogger' in nl) else 'JACKET' if 'jacket' in nl else 'KNIT TOP'
    return r

def extract_sketch_from_bytes(pdf_bytes):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    for pnum in [0, 1]:
        if pnum >= doc.page_count: continue
        imgs = doc[pnum].get_images(full=True)
        for img_info in imgs:
            xref = img_info[0]
            pix = fitz.Pixmap(doc, xref)
            if pix.width > pix.height:
                return PILImage.open(io.BytesIO(pix.tobytes("png")))
    best_pix, best_area = None, 0
    for pnum in range(doc.page_count):
        for img_info in doc[pnum].get_images(full=True):
            xref = img_info[0]
            pix = fitz.Pixmap(doc, xref)
            if pix.width > pix.height and pix.width * pix.height > best_area:
                best_area = pix.width * pix.height
                best_pix = pix
    return PILImage.open(io.BytesIO(best_pix.tobytes("png"))) if best_pix else None

def find_last_subtotal_row(ws):
    last = 3
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and 'SUBTOTAL' in str(val):
            last = r
    return last

def copy_merge_pattern(ws, ref_row, new_row):
    offset = new_row - ref_row
    to_add = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == ref_row:
            to_add.append((mr.min_row + offset, mr.min_col, mr.max_row + offset, mr.max_col))
    for (r1, c1, r2, c2) in to_add:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def copy_cell_style(src, dst):
    if src.font:   dst.font   = copy.copy(src.font)
    if src.border: dst.border = copy.copy(src.border)
    dst.number_format = src.number_format
    dst.fill = NO_FILL

def add_row_to_excel(wb, pdf_bytes, form):
    ws = wb.active
    ref_row = find_last_subtotal_row(ws)
    new_row = ref_row + 4
    sub_row = new_row + 2

    today_dt  = datetime.datetime.now()
    today_str = today_dt.strftime('%Y-%m-%d %H:%M')
    season_str = form.get('season', "HOL'26")
    season_div = f"{season_str}\n{form.get('category','KNIT TOP')}\n{form.get('assign','Hansoll assigned')}"

    for block_offset in range(4):
        src_r = ref_row + block_offset
        dst_r = new_row + block_offset
        ws.row_dimensions[dst_r].height = ws.row_dimensions[src_r].height or 29.0
        for col_idx in range(1, 15):
            try:
                copy_cell_style(ws.cell(row=src_r, column=col_idx), ws.cell(row=dst_r, column=col_idx))
            except: pass

    copy_merge_pattern(ws, ref_row, new_row)

    for r in range(new_row, new_row + 4):
        for col_idx in range(1, 15):
            c = ws.cell(row=r, column=col_idx)
            if type(c).__name__ != 'MergedCell':
                c.fill = YELLOW_FILL

    ws[f'A{new_row}'] = f'=SUBTOTAL(103,D$4:D{new_row})'
    ws[f'A{new_row}'].alignment = CENTER_ALIGN
    ws[f'B{new_row}'] = today_dt
    ws[f'B{new_row}'].number_format = 'm"/"d;@'
    ws[f'B{new_row}'].alignment = CENTER_ALIGN
    ws[f'C{new_row}'] = season_div
    ws[f'C{new_row}'].alignment = CENTER_ALIGN

    if form.get('style'):
        ws[f'D{new_row}'] = int(form['style'])
        ws[f'D{new_row}'].alignment = CENTER_ALIGN

    ws[f'F{new_row}'] = form.get('name', '')
    ws[f'F{new_row}'].alignment = CENTER_ALIGN
    ws[f'G{new_row}'] = form.get('fabric', '')
    ws[f'G{new_row}'].alignment = Alignment(wrap_text=True, vertical='center')

    if form.get('cmph_k'):
        ws[f'K{new_row}'] = float(form['cmph_k'])
        ws[f'K{new_row}'].alignment = CENTER_ALIGN
    if form.get('cmph_l'):
        ws[f'L{new_row}'] = float(form['cmph_l'])
        ws[f'L{new_row}'].alignment = CENTER_ALIGN
    if form.get('cmph_m'):
        ws[f'M{new_row}'] = float(form['cmph_m'])
        ws[f'M{new_row}'].alignment = CENTER_ALIGN

    ws[f'N{new_row}'] = f"★ ADDED {today_str}"
    ws[f'N{new_row}'].alignment = Alignment(wrap_text=True, vertical='center')

    sub_enabled = form.get('sub_enabled')
    if isinstance(sub_enabled, str):
        sub_enabled = sub_enabled.lower() == 'true'

    if not sub_enabled:
        for r_off in [2, 3]:
            for col_idx in range(1, 15):
                c = ws.cell(row=new_row + r_off, column=col_idx)
                if type(c).__name__ != 'MergedCell':
                    b = c.border
                    c.border = Border(
                        left=copy.copy(b.left) if b.left else Side(),
                        right=copy.copy(b.right) if b.right else Side(),
                        top=Side(),
                        bottom=Side()
                    )
    else:
        ws[f'H{sub_row}'] = form.get('sub_qty', '')
        ws[f'H{sub_row}'].alignment = CENTER_ALIGN
        if form.get('sub_cmph'):
            ws[f'M{sub_row}'] = float(form['sub_cmph'])
            ws[f'M{sub_row}'].alignment = CENTER_ALIGN
        if form.get('sub_l'):
            ws[f'L{sub_row}'] = float(form['sub_l'])
            ws[f'L{sub_row}'].alignment = CENTER_ALIGN

    sketch_pil = extract_sketch_from_bytes(pdf_bytes)
    if sketch_pil:
        sketch_resized = sketch_pil.resize((172, 100), PILImage.LANCZOS)
        buf = io.BytesIO()
        sketch_resized.save(buf, format='PNG')
        buf.seek(0)
        xl_img = XLImage(buf)
        xl_img.width  = 172
        xl_img.height = 100
        row_0 = new_row - 1
        anchor = TwoCellAnchor()
        anchor.editAs = 'oneCell'
        anchor._from  = AnchorMarker(col=4, row=row_0,     colOff=916782,  rowOff=88627)
        anchor.to     = AnchorMarker(col=4, row=row_0 + 3, colOff=2782888, rowOff=325711)
        xl_img.anchor = anchor
        ws.add_image(xl_img)

    return new_row

@app.route('/api/python/parse', methods=['POST'])
def parse_pdf():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    pdf_bytes = file.read()
    try:
        data = parse_pdf_text_from_bytes(pdf_bytes)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/python/generate', methods=['POST'])
def generate_excel():
    try:
        if 'excel' not in request.files:
            return jsonify({"error": "No excel file"}), 400
            
        excel_file = request.files['excel']
        wb = load_workbook(io.BytesIO(excel_file.read()))

        # For simplicity, we assume one PDF uploaded as 'pdf'
        # In a multi-PDF scenario, we'd loop over request.files.getlist('pdf')
        pdfs = request.files.getlist('pdf')
        if not pdfs:
            return jsonify({"error": "No pdf files"}), 400

        # Form fields
        form = {
            'style': request.form.get('style', ''),
            'name': request.form.get('name', ''),
            'fabric': request.form.get('fabric', ''),
            'season': request.form.get('season', "HOL'26"),
            'category': request.form.get('category', 'KNIT TOP'),
            'assign': request.form.get('assign', 'Hansoll assigned'),
            'cmph_k': request.form.get('cmph_k', ''),
            'cmph_l': request.form.get('cmph_l', ''),
            'cmph_m': request.form.get('cmph_m', ''),
            'sub_enabled': request.form.get('sub_enabled', 'false'),
            'sub_qty': request.form.get('sub_qty', ''),
            'sub_cmph': request.form.get('sub_cmph', ''),
            'sub_l': request.form.get('sub_l', '')
        }

        for i, pdf_file in enumerate(pdfs):
            pdf_bytes = pdf_file.read()
            if i == 0:
                row_form = dict(form)
            else:
                d = parse_pdf_text_from_bytes(pdf_bytes)
                row_form = {
                    'style': d.get('style',''), 
                    'name': d.get('name',''),
                    'fabric': d.get('fabric',''),
                    'season': form['season'],
                    'category': d.get('category','KNIT TOP'),
                    'assign': form.get('assign','Hansoll assigned'),
                    'cmph_k':'', 'cmph_l':'', 'cmph_m':'',
                    'sub_enabled': False,
                    'sub_qty':'', 'sub_cmph':'', 'sub_remark':''
                }
            add_row_to_excel(wb, pdf_bytes, row_form)

        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        
        return send_file(
            out_buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='CMPH_Recap_Updated.xlsx'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(port=5328)
